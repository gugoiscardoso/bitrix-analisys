# -*- coding: utf-8 -*-
"""
Gera a saída oficial do relatório a partir do store canônico.

    python pipeline/gerar_relatorio.py --de 2026-05-01 --ate 2026-08-05

Produz exatamente dois arquivos em report/:
    relatorio_executivo_<de>_<ate>.xlsx   frentes e subgrupos com métricas (sem plano de ação)
    base_unificada_<de>_<ate>.xlsx        chamados e conversas classificados, linha a linha

Janela pura: tudo — contagens, médias, última incidência, recência — considera apenas
registros com data dentro de [de, ate]. A régua de recência é relativa ao 'ate', não à
data de execução, para que o mesmo período gere sempre o mesmo relatório.
"""
from __future__ import annotations
import argparse
import collections
import json
import statistics
import os
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

RAIZ = Path(__file__).resolve().parent.parent
STORE = RAIZ / "data" / "store"
REPORT = RAIZ / "report"

HDR = PatternFill("solid", fgColor="1F4E79")
HF = Font(color="FFFFFF", bold=True, size=11)
GRP = PatternFill("solid", fgColor="EAF0F6")
CHAM = PatternFill("solid", fgColor="2E6B3E")
CONV = PatternFill("solid", fgColor="6B4E8F")
CALOR = {"recente": (PatternFill("solid", fgColor="F8D7D3"), Font(color="9C2119", bold=True)),
         "medio":   (PatternFill("solid", fgColor="FBE3C8"), Font(color="8A4E10", bold=True)),
         "antigo":  (PatternFill("solid", fgColor="F6EFC6"), Font(color="6B5C09", bold=True))}
BORDA = Border(bottom=Side(style="thin", color="D5DEE6"))

sys.stdout.reconfigure(encoding="utf-8")


def ler(p: Path):
    return json.loads(p.read_text(encoding="utf-8-sig"))


def dias(a: str | None, b: str | None) -> int | None:
    if not a or not b:
        return None
    try:
        d = (datetime.fromisoformat(b).date() - datetime.fromisoformat(a).date()).days
    except (ValueError, TypeError):
        return None
    return d if d >= 0 else None


def media_fmt(vals: list[int]) -> str:
    if not vals:
        return "—"
    m = sum(vals) / len(vals)
    return f"{m:.0f} d  ({len(vals)} chamado{'s' if len(vals) != 1 else ''})"


def calor_de(ultima: str, ate: str) -> tuple[str, str]:
    idade = (datetime.fromisoformat(ate).date() - datetime.fromisoformat(ultima).date()).days
    if idade <= 30:
        return "recente", "Ativa"
    return ("medio", "Atenção") if idade <= 60 else ("antigo", "Antiga")


def carregar(de: str, ate: str):
    cache = ler(STORE / "classificacao.json")
    tax = ler(STORE / "taxonomia.json")
    subs_por_frente = collections.defaultdict(set)
    for s in tax["subgrupos"]:
        subs_por_frente[s["frente"]].add(s["nome"])
    registros = []
    with (STORE / "base_historica.jsonl").open(encoding="utf-8") as fh:
        for linha in fh:
            r = json.loads(linha)
            if not r["fiscal"] or not (de <= r["data"] <= ate):
                continue
            chave = "chamados" if r["tipo"] == "chamado" else "conversas"
            c = cache[chave].get(str(r["id"]))
            if not c:
                continue
            sub = c.get("subgrupo", "")
            # Frentes de tema único: o subgrupo é o próprio tema. Derivado da taxonomia
            # (o tema consta como subgrupo daquela frente) em vez de "== P9" fixo, que
            # deixaria de valer assim que P9 fosse dissolvida.
            if not sub and c["tema"] in subs_por_frente.get(c.get("frente"), ()):
                sub = c["tema"]
            r |= {"tema": c["tema"], "subgrupo": sub,
                  "frente": c.get("frente"), "fonte_cls": c.get("fonte", "")}
            registros.append(r)
    return registros, tax


# ─────────────────────── relatório executivo ───────────────────────

def executivo(regs: list[dict], tax: dict, de: str, ate: str, saida: Path):
    titulos = {f["tag"]: f["titulo"] for f in tax["frentes"]}
    desc = {s["nome"]: s.get("descricao", "") for s in tax["subgrupos"]}

    # agrega por (frente, subgrupo) — sem subgrupo agrupa como "(sem subgrupo atribuído)"
    grupos = collections.defaultdict(lambda: {"ch": 0, "cv": 0, "datas": [],
                                              "total": [], "tech": []})
    for r in regs:
        if not r["frente"]:
            continue
        k = (r["frente"], r["subgrupo"] or "(sem subgrupo atribuído)")
        g = grupos[k]
        g["datas"].append(r["data"])
        if r["tipo"] == "chamado":
            g["ch"] += 1
            if r["status"] == "concluída":
                if (d := dias(r.get("criado_em"), r.get("fechado_em"))) is not None:
                    g["total"].append(d)
                if (d := dias(r.get("inicio_dev"), r.get("alterado_em"))) is not None:
                    g["tech"].append(d)
        else:
            g["cv"] += 1

    carga_frente = collections.Counter()
    for (fr, _), g in grupos.items():
        carga_frente[fr] += g["ch"] + g["cv"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório"
    COLS = ["Grupo principal", "Subgrupo", "Chamados", "Conversas", "Peso no grupo",
            "Tempo total do chamado", "Tempo tech (desenvolvimento)",
            "Última incidência", "Recência"]
    ws.append(COLS)
    for c in ws[1]:
        c.fill, c.font = HDR, HF
        c.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    # Ordem curada de apresentação. Qualquer frente da taxonomia que não esteja aqui
    # entra no fim em vez de SUMIR: com a lista fixa, criar uma frente nova descartava
    # as linhas dela do relatório sem erro nenhum (P10–P13 perderam 1.107 registros
    # assim, na dissolução de P9).
    ordem = ["P1", "P2", "P3", "P4", "P5", "P6", "P9", "P7", "P8"]
    ordem += [f["tag"] for f in tax["frentes"] if f["tag"] not in ordem]
    anterior = None
    for fr in ordem:
        itens = sorted(((k, v) for k, v in grupos.items() if k[0] == fr),
                       key=lambda x: -(x[1]["ch"] + x[1]["cv"]))
        for (_, nome), g in itens:
            carga = g["ch"] + g["cv"]
            pct = round(carga / carga_frente[fr] * 100) if carga_frente[fr] else 0
            ultima = max(g["datas"])
            heat, rot = calor_de(ultima, ate)
            ws.append([f"{fr} — {titulos.get(fr, '')}", nome, g["ch"], g["cv"],
                       "█" * min(max(1, round(pct / 4)), 12) + f"  {pct}%",
                       media_fmt(g["total"]), media_fmt(g["tech"]),
                       datetime.fromisoformat(ultima).strftime("%d/%m/%Y"), rot])
            i = ws.max_row
            if ws.cell(row=i, column=1).value != anterior:
                for c in range(1, len(COLS) + 1):
                    ws.cell(row=i, column=c).fill = GRP
                ws.cell(row=i, column=1).font = Font(bold=True)
                anterior = ws.cell(row=i, column=1).value
            for c in (8, 9):
                ws.cell(row=i, column=c).fill, ws.cell(row=i, column=c).font = CALOR[heat]
                ws.cell(row=i, column=c).alignment = Alignment(horizontal="center", vertical="top")
            for c in (3, 4, 6, 7):
                ws.cell(row=i, column=c).alignment = Alignment(horizontal="center", vertical="top")
            for c in (1, 2):
                ws.cell(row=i, column=c).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(row=i, column=5).font = Font(
                name="Consolas", size=10, bold=pct >= 20,
                color="1F4E79" if pct >= 20 else ("3D83B8" if pct >= 10 else "8FB4D0"))
            for c in range(1, len(COLS) + 1):
                ws.cell(row=i, column=c).border = BORDA
            ws.row_dimensions[i].height = 30

    for i, w in enumerate([32, 46, 11, 11, 17, 19, 19, 14, 11], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "C2"
    n = ws.max_row
    ws.auto_filter.ref = f"A1:I{n}"
    ws.conditional_formatting.add(f"C2:C{n}", DataBarRule(
        start_type="num", start_value=0, end_type="max", color="2E6B3E", showValue=True))
    ws.conditional_formatting.add(f"D2:D{n}", DataBarRule(
        start_type="num", start_value=0, end_type="max", color="6B4E8F", showValue=True))

    legenda(wb, tax, de, ate, regs)
    wb.save(saida)
    return n - 1


def legenda(wb: Workbook, tax: dict, de: str, ate: str, regs: list[dict]):
    ws = wb.create_sheet("Legenda")
    ch = sum(1 for r in regs if r["tipo"] == "chamado")
    cv = len(regs) - ch
    itens = [
        ("Período", f"{de} a {ate}. Janela pura: contagens, médias, última incidência e recência "
                    "consideram apenas registros dentro do período. A régua de recência é relativa "
                    "ao fim da janela, não à data de execução — por isso o mesmo período gera "
                    "sempre o mesmo relatório."),
        ("Volume no período", f"{len(regs)} registros fiscais classificados: {ch} chamados e {cv} conversas."),
        ("Grupo principal", "Frente de trabalho (P1 a P9), da taxonomia versionada."),
        ("Subgrupo", "Subdivisão acionável dentro da frente. '(sem subgrupo atribuído)' reúne o que "
                     "tem tema mas não coube em nenhum subgrupo existente — é o sinal de que a "
                     "taxonomia precisa crescer."),
        ("Chamados / Conversas", "Contagem por canal dentro da janela."),
        ("Peso no grupo", "Participação do subgrupo na carga total da frente (chamados + conversas). "
                          "As linhas de cada frente estão ordenadas por essa carga. Não compare o "
                          "peso entre frentes diferentes."),
        ("Tempo total do chamado  (createdDate → closedDate)",
         "Ciclo completo: média de dias corridos entre a abertura e o fechamento, só dos chamados "
         "concluídos. O número entre parênteses é quantos entraram na média — leia com cuidado "
         "quando forem poucos. É tempo de calendário, não esforço: inclui espera por terceiros."),
        ("Tempo tech (desenvolvimento)  (dateStart → changedDate)",
         "Tempo de execução: entre o início do desenvolvimento e a última alteração do chamado. "
         "Base menor que a coluna anterior (dateStart falta em ~35% dos chamados), portanto NÃO "
         "subtraia uma da outra para obter fila — são conjuntos distintos."),
        ("Última incidência / Recência",
         "Data do registro mais recente do subgrupo no período, considerando chamados E conversas. "
         "Vermelho: últimos 30 dias antes do fim da janela. Laranja: 30 a 60. Amarelo: mais de 60."),
        ("Campos do Bitrix", "createdDate = abertura · dateStart = início do desenvolvimento · "
                             "changedDate = última modificação · closedDate = fechamento. "
                             "timeSpentInLogs não é usado: está vazio porque o controle de tempo "
                             "do Bitrix está desligado, então todas as métricas são de calendário."),
        ("Consistência", f"Classificação vem do cache em data/store/classificacao.json sob a "
                         f"taxonomia {tax['versao']}. Registros já classificados não são reprocessados, "
                         "então rodar o mesmo período duas vezes dá o mesmo resultado."),
        ("Plano de ação", "Não faz parte deste relatório. É artefato curado, mantido à parte em docs/."),
    ]
    for k, v in itens:
        ws.append([k, v])
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 120
    for row in ws.iter_rows():
        row[0].font = Font(bold=True)
        row[1].alignment = Alignment(wrap_text=True, vertical="top")


# ─────────────────────── base unificada ───────────────────────

def base_unificada(regs: list[dict], de: str, ate: str, saida: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Dados"
    COLS = ["Tipo", "Id", "Data", "Mes", "Canal", "Frente", "Tema", "Subgrupo",
            "Status", "TempoTotalDias", "TempoTechDias", "FechadoEm",
            "Cliente", "CNPJ", "DuracaoMin", "Mensagens", "Operador", "Texto"]
    SO_CH = {"Status", "TempoTotalDias", "TempoTechDias", "FechadoEm", "Cliente", "CNPJ"}
    SO_CV = {"DuracaoMin", "Mensagens", "Operador"}
    ws.append(COLS)
    for i, c in enumerate(COLS, 1):
        cel = ws.cell(row=1, column=i)
        cel.font = HF
        cel.fill = CHAM if c in SO_CH else (CONV if c in SO_CV else HDR)

    for r in sorted(regs, key=lambda x: (x["data"], x["tipo"]), reverse=True):
        ch = r["tipo"] == "chamado"
        ws.append([
            "Chamado" if ch else "Conversa", r["id"], r["data"], r["data"][:7], r["canal"],
            r["frente"] or "—", r["tema"], r["subgrupo"],
            r["status"] if ch else "",
            dias(r.get("criado_em"), r.get("fechado_em")) if ch else "",
            dias(r.get("inicio_dev"), r.get("alterado_em")) if ch else "",
            (r.get("fechado_em") or "")[:10] if ch else "",
            r.get("cliente", "") if ch else "", r.get("cnpj", "") if ch else "",
            r.get("duracao_min", "") if not ch else "",
            r.get("mensagens", "") if not ch else "",
            r.get("operador", "") if not ch else "",
            (r.get("titulo") or r.get("texto", ""))[:900],
        ])
    larg = {"Tipo": 10, "Id": 10, "Data": 11, "Mes": 9, "Canal": 15, "Frente": 10,
            "Tema": 46, "Subgrupo": 44, "Status": 14, "TempoTotalDias": 11,
            "TempoTechDias": 11, "FechadoEm": 11, "Cliente": 28, "CNPJ": 18,
            "DuracaoMin": 10, "Mensagens": 10, "Operador": 18, "Texto": 70}
    for i, c in enumerate(COLS, 1):
        ws.column_dimensions[get_column_letter(i)].width = larg[c]
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{ws.max_row}"

    # resumo por tema
    ws2 = wb.create_sheet("Resumo por tema")
    ws2.append(["Frente", "Tema", "Chamados", "Conversas", "Total"])
    for c in ws2[1]:
        c.fill, c.font = HDR, HF
    agg = collections.defaultdict(lambda: [0, 0])
    for r in regs:
        agg[(r["frente"] or "—", r["tema"])][0 if r["tipo"] == "chamado" else 1] += 1
    for (fr, tema), (a, b) in sorted(agg.items(), key=lambda x: -(x[1][0] + x[1][1])):
        ws2.append([fr, tema, a, b, a + b])
    for i, w in enumerate([10, 56, 11, 11, 9], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"
    wb.save(saida)


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--de", required=True)
    ap.add_argument("--ate", default=date.today().isoformat())
    ap.add_argument("--abrir", action="store_true",
                    help="abre o relatório executivo no aplicativo padrão. NÃO é o padrão "
                         "de propósito: arquivo aberto no Excel trava a gravação da próxima "
                         "execução com 'Permission denied'.")
    a = ap.parse_args()
    REPORT.mkdir(exist_ok=True)

    regs, tax = carregar(a.de, a.ate)
    if not regs:
        print(f"Nenhum registro fiscal classificado em {a.de}..{a.ate}.")
        return 1

    f1 = REPORT / f"relatorio_executivo_{a.de}_{a.ate}.xlsx"
    f2 = REPORT / f"base_unificada_{a.de}_{a.ate}.xlsx"
    linhas = executivo(regs, tax, a.de, a.ate, f1)
    base_unificada(regs, a.de, a.ate, f2)

    ch = sum(1 for r in regs if r["tipo"] == "chamado")
    sem_sub = sum(1 for r in regs if r["frente"] and not r["subgrupo"])
    print(f"Janela {a.de} a {a.ate}")
    print(f"  {len(regs)} registros fiscais ({ch} chamados, {len(regs)-ch} conversas)")
    print(f"  {linhas} linhas de subgrupo no relatório")
    print(f"  {sem_sub} sem subgrupo atribuído ({sem_sub/len(regs)*100:.1f}%)")
    # Caminho ABSOLUTO. O relativo economizava caracteres e fazia quem não estava no
    # diretório do projeto ter de adivinhar onde o arquivo foi parar.
    print(f"\n  {f1.resolve()}")
    print(f"  {f2.resolve()}")

    if a.abrir:
        import subprocess
        try:
            if sys.platform == "win32":
                os.startfile(str(f1.resolve()))            # noqa: S606
            else:
                subprocess.run(["xdg-open", str(f1.resolve())], check=False)
            print("\n  Aberto no aplicativo padrão. Feche antes da próxima execução —")
            print("  arquivo aberto no Excel bloqueia a gravação.")
        except Exception as e:
            print(f"\n  Não consegui abrir ({e.__class__.__name__}). O caminho está acima.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
