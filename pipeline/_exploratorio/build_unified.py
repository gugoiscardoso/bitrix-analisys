# -*- coding: utf-8 -*-
"""Planilha unificada: chamados + conversas no mesmo tabelão, com colunas por tipo."""
import json, re, sys, io, collections
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

SCRATCH = r"C:\Users\Lucas\AppData\Local\Temp\claude\C--Dev-Qigger-Ultracar-bitrix-analisys\a9f3de17-99fb-4920-8408-48ea262b563e\scratchpad"
OUT = r"C:\Dev\Qigger\Ultracar\bitrix-analisys\output\base_unificada_fiscal_2026-08-05.xlsx"
EXPORT = r"C:\Dev\Qigger\Ultracar\bitrix-analisys\output\bitrix_export_139_20260804_200226.json"

# ---------- mapa tema -> proposta ----------
PROP = {
    "Rejeição/erro de validação (schema XML, E0xxx, tags, campos, IE)": "P1 Pré-validação",
    "Status dessincronizado sistema x prefeitura/SEFAZ": "P2 Reconciliação",
    "Nota travada em processamento/transmissão sem retorno": "P2 Reconciliação",
    "Numeração/duplicidade (pulos, RPS, DPS, inutilização)": "P2 Reconciliação",
    "Cálculo/exibição de impostos errada (PIS/COFINS/ICMS/ISS/IBS/CBS/retenções)": "P3 Impostos",
    "PDF/DANFE/impressão (não gera, dados fora do lugar)": "P4 PDF/DANFE",
    "Sem causa identificável na descrição": "P5 Triagem",
    "Particularidade municipal de NFS-e (layout, homologação, instabilidade da prefeitura)": "P6 Municipal+Certificado",
    "Certificado digital (cadastro, vencimento, atualização)": "P6 Municipal+Certificado",
    "Nota de devolução/garantia/remessa/complemento": "P7 Devolução guiada",
    "Dúvida de uso / orientação (how-to fiscal)": "P8 Autonomia",
    "Configuração assistida pelo suporte": "P8 Autonomia",
    "Acompanhamento de chamado já aberto": "P8 Autonomia",
    "Instabilidade geral / lentidão do sistema": "P8 Autonomia",
    "Cadastro/config fiscal (NCM, CFOP, CST, cód. serviço, SPED, regime)": "P9 Frentes complementares",
    "Integração financeiro/estoque/OS com a nota": "P9 Frentes complementares",
    "Cancelamento/exclusão de nota": "P9 Frentes complementares",
    "Relatórios fiscais divergentes": "P9 Frentes complementares",
    "XML de compra / importação / manifestação do destinatário": "P9 Frentes complementares",
    "Envio de nota por e-mail falha": "P9 Frentes complementares",
    "NFS-e interna / API interna": "P9 Frentes complementares",
    "Conversa vazia / sem conteúdo útil": "— (descartada)",
    "Não fiscal (falso positivo)": "— (descartada)",
}
NORM = {
    "Certificado digital": "Certificado digital (cadastro, vencimento, atualização)",
    "Particularidade municipal de NFS-e (layout, homologação, prefeitura)":
        "Particularidade municipal de NFS-e (layout, homologação, instabilidade da prefeitura)",
}
STATUS = {"1": "nova", "2": "pendente", "3": "em andamento", "4": "aguard. controle",
          "5": "concluída", "6": "adiada", "7": "recusada"}
CNPJ_RE = re.compile(r"\b(\d{2}\.?\d{3}\.?\d{3}/?\d{4}-?\d{2})\b")

def strip_bb(s):
    return re.sub(r"\[/?[A-Za-z][^\]]*\]", "", s or "")

def extract_client(t):
    m = re.search(r"cliente\s*:?\s*(.+?)(?:\s*[-–]\s*(?:id|cnpj)|\s*cnpj|$)", t, re.IGNORECASE)
    if m:
        c = m.group(1).strip(" -–:|")
        if 2 < len(c) < 90 and not CNPJ_RE.match(c):
            return c
    return ""

def extract_cnpj(t):
    m = CNPJ_RE.search(t.replace(" ", ""))
    if m:
        return m.group(1)
    m = re.search(r"\[filial:\s*(\d{14})\]", t, re.IGNORECASE)
    return m.group(1) if m else ""

def days(a, b):
    if not a or not b:
        return ""
    try:
        return (datetime.fromisoformat(b).date() - datetime.fromisoformat(a).date()).days
    except ValueError:
        return ""

# subgrupo por id de conversa (censo P7/P8)
sub_by_sid = {}
_sc = json.load(open(SCRATCH + r"\sub_conv_final.json", encoding="utf-8-sig"))
for _fr in _sc.values():
    for _k, _v in _fr.items():
        sub_by_sid[int(_k)] = ("" if _v.startswith("Outro") else _v)

# subgrupo por id de chamado
sub_by_id = {}
P = json.load(open(SCRATCH + r"\propostas.json", encoding="utf-8"))
for p in P.values():
    for sg in p["subgrupos"]:
        for i in sg["ids"]:
            sub_by_id[i] = sg["nome"]

rows = []

# ---------- CHAMADOS ----------
export = json.load(open(EXPORT, encoding="utf-8"))
by_id = {t["task"]["id"]: t["task"] for t in export["tasks"]}
v2 = json.load(open(SCRATCH + r"\class_final_v2.json", encoding="utf-8"))
final, fonte = v2["final"], v2["fonte"]
id2problem = {}
for g, items in json.load(open(SCRATCH + r"\fiscal_final.json", encoding="utf-8")).items():
    for i in items:
        id2problem[i["id"]] = i["problem"]

for tid in sorted(final, key=int):
    t = by_id[tid]
    title = t.get("title") or ""
    cls = NORM.get(final[tid], final[tid])
    created = (t.get("createdDate") or "")[:10]
    rows.append({
        "Tipo": "Chamado", "Id": int(tid),
        "Link": f"https://ultracar.bitrix24.com/workgroups/group/139/tasks/task/view/{tid}/",
        "Data": created, "Mes": created[:7],
        "Canal": "Matrix (bot)" if re.search(r"\[matrix", title, re.IGNORECASE) else "Analista",
        "Classificacao": cls, "Subgrupo": sub_by_id.get(tid, ""),
        "Proposta": PROP.get(cls, "—"), "FonteClassificacao": fonte.get(tid, ""),
        "Status": STATUS.get(str(t.get("status")), ""),
        "TempoResolucaoDias": days(t.get("createdDate"), t.get("closedDate")),
        "FechadoEm": (t.get("closedDate") or "")[:10],
        "Cliente": extract_client(title), "CNPJ": extract_cnpj(title),
        "PrefixoOriginal": (re.match(r"\s*\[([^\]]+)\]", title).group(1).strip()
                            if re.match(r"\s*\[([^\]]+)\]", title) else ""),
        "TemChamadoVinculado": "", "DuracaoMin": "", "TotalMensagens": "", "Operador": "",
        "Resumo": re.sub(r"\s+", " ", id2problem.get(tid, ""))[:450],
        "TituloOriginal": re.sub(r"\s+", " ", title)[:400],
        "TextoCompleto": strip_bb(t.get("description") or "").strip()[:4000],
    })
print(f"Chamados: {len(rows)}")

# ---------- CONVERSAS ----------
conv = json.load(open(SCRATCH + r"\conv_fiscais.json", encoding="utf-8"))
cls_conv = {}
for i in range(1, 17):
    for k, v in json.load(open(SCRATCH + rf"\class_conv_{i}.json", encoding="utf-8")).items():
        cls_conv[int(k)] = v

n_conv = 0
for c in conv:
    sid = int(c["SessionId"])
    tema = cls_conv.get(sid)
    if tema is None:
        continue
    started = str(c.get("StartedAt") or "")[:10]
    rows.append({
        "Tipo": "Conversa", "Id": sid, "Link": "",
        "Data": started, "Mes": started[:7],
        "Canal": c.get("Channel") or "",
        "Classificacao": tema, "Subgrupo": sub_by_sid.get(sid, ""),
        "Proposta": PROP.get(tema, "—"), "FonteClassificacao": "chat (LLM)",
        "Status": "", "TempoResolucaoDias": "", "FechadoEm": "",
        "Cliente": "", "CNPJ": "", "PrefixoOriginal": "",
        "TemChamadoVinculado": "Sim" if c.get("tem_chamado") else "Não",
        "DuracaoMin": round(c.get("DurationMinutes") or 0),
        "TotalMensagens": int(c.get("TotalMessages") or 0),
        "Operador": c.get("OperatorName") or "",
        "Resumo": re.sub(r"\s+", " ", str(c.get("digest") or ""))[:450],
        "TituloOriginal": "", "TextoCompleto": re.sub(r"\s+", " ", str(c.get("digest") or ""))[:4000],
    })
    n_conv += 1
print(f"Conversas: {n_conv} | TOTAL linhas: {len(rows)}")

# ---------- Excel ----------
wb = Workbook()
HDR = PatternFill("solid", fgColor="1F4E79")
HF = Font(color="FFFFFF", bold=True)
CHAM = PatternFill("solid", fgColor="2E6B3E")   # verde: colunas só de chamado
CONV = PatternFill("solid", fgColor="6B4E8F")   # roxo: colunas só de conversa

ws = wb.active
ws.title = "Dados"
cols = list(rows[0].keys())
SO_CHAMADO = {"Link", "Status", "TempoResolucaoDias", "FechadoEm", "Cliente", "CNPJ",
              "PrefixoOriginal", "TituloOriginal", "Subgrupo"}
SO_CONVERSA = {"TemChamadoVinculado", "DuracaoMin", "TotalMensagens", "Operador"}
ws.append(cols)
for i, c in enumerate(cols, 1):
    cell = ws.cell(row=1, column=i)
    cell.font = HF
    cell.fill = CHAM if c in SO_CHAMADO else (CONV if c in SO_CONVERSA else HDR)
for r in rows:
    ws.append([r[c] for c in cols])
lk = cols.index("Link") + 1
for i in range(2, len(rows) + 2):
    cell = ws.cell(row=i, column=lk)
    if cell.value:
        cell.hyperlink = cell.value
        cell.value = "abrir"
        cell.font = Font(color="0563C1", underline="single")
W = {"Tipo": 10, "Id": 10, "Link": 8, "Data": 11, "Mes": 9, "Canal": 15, "Classificacao": 48,
     "Subgrupo": 44, "Proposta": 24, "FonteClassificacao": 17, "Status": 14,
     "TempoResolucaoDias": 10, "FechadoEm": 11, "Cliente": 30, "CNPJ": 18, "PrefixoOriginal": 14,
     "TemChamadoVinculado": 12, "DuracaoMin": 10, "TotalMensagens": 11, "Operador": 20,
     "Resumo": 60, "TituloOriginal": 45, "TextoCompleto": 60}
for i, c in enumerate(cols, 1):
    ws.column_dimensions[get_column_letter(i)].width = W.get(c, 18)
ws.freeze_panes = "C2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(rows)+1}"

# ---------- Resumo cruzado ----------
ws2 = wb.create_sheet("Resumo por tema")
ws2.append(["Classificação", "Proposta", "Chamados", "Conversas", "Total",
            "Iceberg (conv/cham)", "Conversas sem chamado", "% sem chamado"])
for c in ws2[1]:
    c.fill, c.font = HDR, HF
ch = collections.Counter(r["Classificacao"] for r in rows if r["Tipo"] == "Chamado")
cv = collections.Counter(r["Classificacao"] for r in rows if r["Tipo"] == "Conversa")
sem = collections.Counter(r["Classificacao"] for r in rows
                          if r["Tipo"] == "Conversa" and r["TemChamadoVinculado"] == "Não")
for tema in sorted(set(ch) | set(cv), key=lambda t: -(cv.get(t, 0) + ch.get(t, 0))):
    a, b = ch.get(tema, 0), cv.get(tema, 0)
    ws2.append([tema, PROP.get(tema, "—"), a, b, a + b,
                round(b / a, 1) if a else "só chat",
                sem.get(tema, 0), round(sem.get(tema, 0) / b * 100) if b else ""])
ws2.append(["TOTAL", "", sum(ch.values()), sum(cv.values()), sum(ch.values()) + sum(cv.values()),
            round(sum(cv.values()) / sum(ch.values()), 1), sum(sem.values()),
            round(sum(sem.values()) / sum(cv.values()) * 100)])
ws2[f"A{ws2.max_row}"].font = Font(bold=True)
for i, w in enumerate([56, 24, 11, 11, 9, 18, 20, 13], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = "A2"

# ---------- Metodologia ----------
ws3 = wb.create_sheet("Metodologia")
meta = [
    ["O que é esta base", "Base única com as duas fontes de atendimento fiscal da Ultracar: chamados do Bitrix (grupo 139) e conversas de Open Lines (WhatsApp/LiveChat/Telegram). A coluna Tipo separa as duas."],
    ["Período", "2026-05-01 a 2026-08-05"],
    ["Colunas comuns", "Tipo, Id, Data, Mes, Canal, Classificacao, Proposta, FonteClassificacao, Resumo, TextoCompleto"],
    ["Colunas só de Chamado (cabeçalho verde)", "Link, Subgrupo, Status, TempoResolucaoDias, FechadoEm, Cliente, CNPJ, PrefixoOriginal, TituloOriginal"],
    ["Colunas só de Conversa (cabeçalho roxo)", "TemChamadoVinculado, DuracaoMin, TotalMensagens, Operador"],
    ["Classificação", "Taxonomia única de 16 temas fiscais + 6 categorias específicas de chat (how-to, config assistida, acompanhamento, instabilidade, vazia, falso positivo). Toda a classificação foi feita por leitura completa do texto (LLM), em lotes."],
    ["Subgrupo", "Subdivisão acionável dentro da frente. Preenchido nos 370 chamados das frentes P1 a P6, cujos 41 subgrupos foram derivados lendo chamado a chamado. Preenchido tambem nas 3.264 conversas de P1, P7 e P8, classificadas uma a uma. Ainda VAZIO para as conversas de P2 a P6 e P9 e para os chamados de P7/P9 — nesses casos os subgrupos foram derivados por amostragem (P7/P8) ou ainda não foram derivados (P9), então não há atribuição individual confiável. Use a coluna Proposta para agrupar esses registros."],
    ["Iceberg", "Conversas dividido por chamados do mesmo tema. Alto = dor absorvida no balcão que quase nunca vira ticket. Global: 11,4x."],
    ["Ressalva sobre DuracaoMin", "É a janela da sessão (abertura até fechamento/auto-close), inclui ociosidade e espera; NÃO é tempo de atendimento efetivo. Some com cautela — o total excede a capacidade da equipe. Para peso relativo entre temas, prefira TotalMensagens."],
    ["Fontes", "bitrix_export_139_20260804_200226.json e conversations_export_20260805_134927.xlsx"],
    ["Relatórios", "analise_chamados_fiscais_2026-08-04.md (chamados) e analise_balcao_conversas_2026-08-05.md (balcão)"],
    ["Gerado em", "2026-08-05"],
]
for r in meta:
    ws3.append(r)
ws3.column_dimensions["A"].width = 34
ws3.column_dimensions["B"].width = 120
for r in ws3.iter_rows():
    r[0].font = Font(bold=True)
    r[1].alignment = Alignment(wrap_text=True, vertical="top")

wb.save(OUT)
print(f"Salvo: {OUT}")
