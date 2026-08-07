# -*- coding: utf-8 -*-
"""Planilha executiva do plano de ação: grupo > subgrupo > última incidência > ações."""
import json, sys, io, re
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
sys.path.insert(0, r"C:\Users\Lucas\AppData\Local\Temp\claude\C--Dev-Qigger-Ultracar-bitrix-analisys\a9f3de17-99fb-4920-8408-48ea262b563e\scratchpad")
from acoes import ACOES, P9

SCRATCH = r"C:\Users\Lucas\AppData\Local\Temp\claude\C--Dev-Qigger-Ultracar-bitrix-analisys\a9f3de17-99fb-4920-8408-48ea262b563e\scratchpad"
OUT = r"C:\Dev\Qigger\Ultracar\bitrix-analisys\output\plano_acao_executivo_2026-08-05.xlsx"
REF = date(2026, 8, 5)
ACOES_USER = json.load(open(SCRATCH + r"\acoes_user.json", encoding="utf-8-sig"))
TEMPOS = json.load(open(SCRATCH + r"\tempo_medio.json", encoding="utf-8-sig"))
TECH = json.load(open(SCRATCH + r"\tempo_tech.json", encoding="utf-8-sig"))

def tech_de(tag, nome):
    """Media de dias entre o inicio do desenvolvimento (dateStart) e a ultima
    alteracao do chamado (changedDate), nos chamados concluidos."""
    ds = TECH.get(f"{tag}||{nome}")
    if not ds:
        return "—"
    m = sum(ds) / len(ds)
    return f"{m:.0f} d  ({len(ds)} chamado{'s' if len(ds) != 1 else ''})"

def tempo_de(tag, nome):
    """Media de dias ate o fechamento, sobre os chamados CONCLUIDOS do subgrupo."""
    ds = TEMPOS.get(f"{tag}||{nome}")
    if not ds:
        return "—"
    m = sum(ds) / len(ds)
    return f"{m:.0f} d  ({len(ds)} chamado{chr(115) if len(ds) != 1 else chr(39)+chr(39)})".replace(chr(39)+chr(39), "")

def plano_de(tag, nome, fallback):
    """Acoes revisadas pelo usuario tem precedencia; fallback e o texto gerado."""
    k = f"{tag}||{nome}"
    src = ACOES_USER[k] if k in ACOES_USER else fallback   # lista vazia = remoção intencional
    return "\n".join(f"{n}. {a}" for n, a in enumerate(src, 1))

P = json.load(open(SCRATCH + r"\propostas.json", encoding="utf-8"))
p7 = json.load(open(SCRATCH + r"\sub_p7.json", encoding="utf-8"))
p8 = json.load(open(SCRATCH + r"\sub_p8.json", encoding="utf-8"))

GRUPO = {
    "prop1": ("P1", "Pré-validação da nota antes da transmissão", 1409, 21.7, "Confirmada nº 1 nos dois canais"),
    "prop2": ("P2", "Job de reconciliação com SEFAZ e prefeituras", 499, 9.0, "Manter — tamanho já era fiel"),
    "prop3": ("P3", "Pacote de conformidade tributária (IBS/CBS)", 377, 7.2, "Manter — cresce com a reforma"),
    "prop4": ("P4", "Robustez de PDF/DANFE e impressão", 132, 2.0, "Rebaixar — iceberg 3,1x"),
    "prop5": ("P5", "Triagem com evidência obrigatória", 0, 0.0, "Manter — problema de processo"),
    "prop6": ("P6", "Homologação municipal + certificado digital", 642, 10.6, "Subir — 2,6x maior no chat"),
}
# Frentes com base em chamados primeiro (ordem de prioridade da análise de chamados),
# depois as que nascem só das conversas — menos prioritárias entre blocos, ordenadas entre si.
ORDEM = ["prop1", "prop2", "prop3", "prop4", "prop5", "prop6", "P9", "P7", "P8"]

def heat_of(iso):
    age = (REF - datetime.fromisoformat(iso).date()).days
    return ("Ativa" if age <= 30 else ("Atenção" if age <= 60 else "Antiga")), age

def find_acoes(key, nome, sg=None):
    for (k, frag), acts in ACOES.items():
        if k == key and frag.lower() in nome.lower():
            return acts
    if sg and sg.get("acoes"):        # subgrupos novos guardam as ações na própria definição
        return sg["acoes"]
    return ["(a definir)"]

rows = []
for g in ORDEM:
    if g in GRUPO:
        tag, titulo, conv, carga, verd = GRUPO[g]
        p = P[g]
        for sg in sorted(p["subgrupos"], key=lambda x: -(x["count"] + (x.get("n_conv") or 0))):
            heat, age = heat_of(sg["ultima"])
            rows.append({
                "Grupo": f"{tag} — {titulo}", "GrupoTag": tag,
                "VolumeGrupo": f"{p['total']} chamados" + (f" · {conv} conversas" if conv else ""),
                "Veredito": verd,
                "Subgrupo": sg["nome"], "Chamados": sg["count"], "Conversas": sg.get("n_conv"), "Tempo": tempo_de(tag, sg["nome"]), "Tech": tech_de(tag, sg["nome"]), "_base": sg["count"] + (sg.get("n_conv") or 0), "_tot": sum(y["count"] + (y.get("n_conv") or 0) for y in p["subgrupos"]),
                "EmAberto": sg["abertos"], "DiasTratativa": sg["soma_dias"],
                "UltimaIncidencia": datetime.fromisoformat(sg["ultima"]).strftime("%d/%m/%Y"),
                "Recencia": heat, "_heat": sg["heat"], "_age": age,
                "Contexto": sg["descricao"],
                "PlanoDeAcao": plano_de(tag, sg["nome"], find_acoes(g, sg["nome"], sg)),
            })
    elif g == "P7":
        for sg in sorted(p7["subgrupos"], key=lambda x: -(x.get("n_exato") or 0)):
            n = sg.get("n_exato") or 0
            rows.append({
                "Grupo": "P7 — Fluxo guiado de devolução, garantia e remessa", "GrupoTag": "P7",
                "VolumeGrupo": "24 chamados · 861 conversas", "Veredito": "CRIAR — iceberg 35,9x, 2ª maior dor real",
                "Subgrupo": sg["nome"], "Chamados": None, "Conversas": n, "Tempo": "—", "Tech": "—", "_base": sg["n_exato"], "_tot": 861 if g=="P7" else 994, "EmAberto": "",
                "DiasTratativa": "", "UltimaIncidencia": datetime.fromisoformat(sg["ultima"]).strftime("%d/%m/%Y"), "Recencia": heat_of(sg["ultima"])[0],
                "_heat": sg["heat"], "_age": sg.get("idade_dias",0), "Contexto": sg["descricao"],
                "PlanoDeAcao": plano_de("P7", sg["nome"], sg["acoes"]),
            })
    elif g == "P8":
        for sg in sorted(p8["subgrupos"], key=lambda x: -x["n_exato"]):
            n = sg.get("n_exato") or 0
            rows.append({
                "Grupo": "P8 — Autonomia do cliente (self-service)", "GrupoTag": "P8",
                "VolumeGrupo": "0 chamados · 994 conversas", "Veredito": "CRIAR — invisível nos tickets",
                "Subgrupo": sg["nome"], "Chamados": None, "Conversas": n, "Tempo": "—", "Tech": "—", "_base": sg["n_exato"], "_tot": 861 if g=="P7" else 994, "EmAberto": "",
                "DiasTratativa": "", "UltimaIncidencia": datetime.fromisoformat(sg["ultima"]).strftime("%d/%m/%Y"), "Recencia": heat_of(sg["ultima"])[0],
                "_heat": sg["heat"], "_age": sg.get("idade_dias",0), "Contexto": sg["descricao"],
                "PlanoDeAcao": plano_de("P8", sg["nome"], sg["acoes"]),
            })
    elif g == "P9":
        for nome, ch, ab, ult, heat, cv, acts in P9:
            h, age = heat_of(ult)
            rows.append({
                "Grupo": "P9 — Frentes complementares", "GrupoTag": "P9",
                "VolumeGrupo": "128 chamados · 1.061 conversas", "Veredito": "Avaliar — fora das 6 propostas originais",
                "Subgrupo": nome, "Chamados": ch, "Conversas": cv, "Tempo": tempo_de("P9", nome), "Tech": tech_de("P9", nome), "_base": ch, "_tot": 128, "EmAberto": ab, "DiasTratativa": "",
                "UltimaIncidencia": datetime.fromisoformat(ult).strftime("%d/%m/%Y"),
                "Recencia": h, "_heat": heat, "_age": age,
                "Contexto": f"{cv} conversas no chat sobre o mesmo tema.",
                "PlanoDeAcao": plano_de("P9", nome, acts),
            })

for r in rows:
    pct = round(r["_base"] / r["_tot"] * 100) if r.get("_tot") else 0
    r["_pct"] = pct
    cheios = max(1, round(pct / 4))          # escala: 25% = 6 blocos
    r["Peso"] = "█" * min(cheios, 12) + f"  {pct}%"

print(f"Linhas no plano: {len(rows)}")

# ---------------- Excel ----------------
wb = Workbook()
HDR = PatternFill("solid", fgColor="1F4E79")
HF = Font(color="FFFFFF", bold=True, size=11)
HEAT_FILL = {"recente": PatternFill("solid", fgColor="F8D7D3"),
             "medio": PatternFill("solid", fgColor="FBE3C8"),
             "antigo": PatternFill("solid", fgColor="F6EFC6")}
HEAT_FONT = {"recente": Font(color="9C2119", bold=True),
             "medio": Font(color="8A4E10", bold=True),
             "antigo": Font(color="6B5C09", bold=True)}
GRP_FILL = PatternFill("solid", fgColor="EAF0F6")
thin = Side(style="thin", color="D5DEE6")
BORDER = Border(bottom=thin)

ws = wb.active
ws.title = "Plano de ação"
COLS = ["Grupo", "Subgrupo", "Chamados", "Conversas", "Peso", "Tempo", "Tech", "UltimaIncidencia", "Recencia", "PlanoDeAcao"]
TITULOS = ["Grupo principal", "Subgrupo", "Chamados", "Conversas", "Peso no grupo", "Tempo total do chamado", "Tempo tech (desenvolvimento)", "Última incidência", "Recência", "Plano de ação"]
ws.append(TITULOS)
for c in ws[1]:
    c.fill, c.font = HDR, HF
    c.alignment = Alignment(vertical="center", wrap_text=True)
ws.row_dimensions[1].height = 30

prev = None
for r in rows:
    ws.append([r[c] for c in COLS])
    i = ws.max_row
    if r["Grupo"] != prev:
        for c in range(1, len(COLS) + 1):
            ws.cell(row=i, column=c).fill = GRP_FILL
        ws.cell(row=i, column=1).font = Font(bold=True)
        prev = r["Grupo"]
    hc = ws.cell(row=i, column=COLS.index("UltimaIncidencia") + 1)
    rc = ws.cell(row=i, column=COLS.index("Recencia") + 1)
    hc.fill = rc.fill = HEAT_FILL[r["_heat"]]
    hc.font = rc.font = HEAT_FONT[r["_heat"]]
    hc.alignment = rc.alignment = Alignment(horizontal="center", vertical="top")
    ws.cell(row=i, column=COLS.index("PlanoDeAcao") + 1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row=i, column=COLS.index("Subgrupo") + 1).alignment = Alignment(wrap_text=True, vertical="top")
    for _c in ("Chamados", "Conversas", "Tempo", "Tech"):
        ws.cell(row=i, column=COLS.index(_c) + 1).alignment = Alignment(horizontal="center", vertical="top")
    pc = ws.cell(row=i, column=COLS.index("Peso") + 1)
    pc.alignment = Alignment(horizontal="left", vertical="top")
    _p = r["_pct"]
    _cor = "1F4E79" if _p >= 20 else ("3D83B8" if _p >= 10 else "8FB4D0")
    pc.font = Font(name="Consolas", size=10, color=_cor, bold=_p >= 20)
    ws.cell(row=i, column=COLS.index("Grupo") + 1).alignment = Alignment(wrap_text=True, vertical="top")
    for c in range(1, len(COLS) + 1):
        ws.cell(row=i, column=c).border = BORDER
    n_acoes = r["PlanoDeAcao"].count("\n") + 1
    ws.row_dimensions[i].height = max(46, min(150, 15 * n_acoes + 14))

WID = {"Grupo": 30, "Subgrupo": 42, "Chamados": 10, "Conversas": 10, "Peso": 16, "Tempo": 19, "Tech": 19,
       "UltimaIncidencia": 13, "Recencia": 10, "PlanoDeAcao": 92}
for i, c in enumerate(COLS, 1):
    ws.column_dimensions[get_column_letter(i)].width = WID[c]
from openpyxl.formatting.rule import DataBarRule
_last = len(rows) + 1
ws.conditional_formatting.add(f"C2:C{_last}",
    DataBarRule(start_type="num", start_value=0, end_type="max", color="2E6B3E", showValue=True))
ws.conditional_formatting.add(f"D2:D{_last}",
    DataBarRule(start_type="num", start_value=0, end_type="max", color="6B4E8F", showValue=True))

ws.freeze_panes = "C2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{len(rows)+1}"

# ---------------- Legenda ----------------
ws3 = wb.create_sheet("Legenda")
leg = [
 ["Como usar", "Uma linha por subgrupo, agrupadas por frente. Ordem: primeiro as frentes originadas dos chamados (P1 a P6, depois P9), na prioridade da análise de chamados; em seguida as frentes que aparecem apenas nas conversas de chat (P7 e P8), ordenadas entre si. As ações são específicas de cada subgrupo."],
 ["Grupo principal", "Frente de trabalho (P1 a P9). P1–P6 nasceram dos 522 chamados; P7 e P8 nasceram da análise das 5.975 conversas de chat; P9 reúne temas fora das propostas originais."],
 ["Subgrupo", "Subdivisão acionável dentro da frente — cada uma implementável de forma independente."],
 ["Tempo total do chamado  (createdDate → closedDate)", "Ciclo completo: média de dias corridos entre a ABERTURA do chamado (campo createdDate do Bitrix) e o seu FECHAMENTO (closedDate). Só entram chamados concluídos; o número entre parênteses é quantos entraram na média. Leia com cuidado quando forem poucos (1 a 3): a média vira anedota. É tempo de calendário, não esforço — inclui espera por terceiros (prefeitura, integradora, contabilidade do cliente). Aparece '—' em P7 e P8: a P8 não tem chamados, e os 22 chamados de devolução da P7 não têm atribuição individual por subgrupo (os subgrupos vieram das conversas); a média da frente P7 inteira é 17 dias."],
 ["Tempo tech (desenvolvimento)  (dateStart → changedDate)", "Tempo de EXECUÇÃO: média de dias entre o INÍCIO do desenvolvimento (campo dateStart do Bitrix) e a ÚLTIMA ALTERAÇÃO do chamado (changedDate), que nos concluídos equivale ao término do trabalho. Só entram chamados concluídos. ATENÇÃO: base diferente da coluna anterior — o tempo tech só existe nos 583 chamados concluídos que têm dateStart preenchido, contra 872 da coluna total. Por isso NÃO subtraia uma coluna da outra para obter tempo de fila: são médias de conjuntos distintos e o resultado pode até ficar negativo. A fila real (createdDate → dateStart) tem mediana 0 e média 2,3 dias — os chamados são pegos rápido; o que demora é resolver."],
 ["Chamados", "Chamados do subgrupo (contagem exata — cada chamado foi lido e atribuído individualmente). Em P7 e P8 aparece '—' ou 0 porque a dor quase não gera chamado: seus subgrupos foram derivados das conversas."],
 ["Peso no grupo", "Participação do subgrupo dentro da própria frente, em barra e percentual, calculada sobre a CARGA TOTAL (chamados + conversas). As linhas de cada frente estão ordenadas por essa carga. Não compare o peso entre frentes diferentes."],
 ["Conversas", "Conversas de chat do subgrupo. Em P1, P7, P8 e P9 é contagem EXATA — as 1.855 conversas de P7 e P8 foram classificadas uma a uma por subgrupo. Em P2 a P6 aparece vazio: ali as conversas estão classificadas por tema, ainda não por subgrupo."],
 ["Última incidência", "Data do registro mais recente do subgrupo, considerando AS DUAS FONTES: chamados e conversas de chat. Até 05/08 esta coluna olhava apenas os chamados, o que envelhecia artificialmente vários subgrupos — 44 deles mudaram de data ao incluir o chat. Vermelho = últimos 30 dias (dor ativa); laranja = 30 a 60 dias; amarelo = mais de 60 dias (possivelmente já corrigido — confirmar com o time antes de investir)."],
 ["Dias tratativa", "Soma de (fechamento − abertura) dos chamados concluídos do subgrupo. Proxy de custo de suporte, em dias de calendário."],
 ["Iceberg", "Conversas de chat divididas por chamados do mesmo tema. Alto = dor absorvida no balcão que quase nunca vira ticket. Global: 11,4x."],
 ["Plano de ação", "Ações concretas propostas, derivadas da leitura dos chamados e das conversas. Numeradas por ordem sugerida de execução dentro do subgrupo."],
 ["\"FASE 1\" nas ações", "Aparece apenas na frente P2 (job de reconciliação). Significa o PRIMEIRO INCREMENTO DE ENTREGA do job — não a frente inteira. A P2 tem 7 funções independentes; entregar todas de uma vez é grande demais, então a Fase 1 recomendada são as duas de maior carga: 'nota presa em processando' (207 conversas) e 'duplicidade de DPS/RPS' (184). Juntas cobrem 71% da frente. As demais funções vêm em incrementos seguintes."],
 ["Funções que compartilham mecanismo (P2)", "Informação para o planejamento técnico, independente da Fase 1: as funções 'nota autorizada no portal continua em digitação', 'nota cancelada na prefeitura permanece autorizada' e 'nota emitida na SEFAZ não aparece' usam a MESMA mecânica — consultar protocolo/chave no provedor e atualizar o estado local. Somam 64 conversas e 57 chamados. Vale construir esse mecanismo uma vez e atender as três de uma vez, mesmo que venham depois da Fase 1."],
 ["Base de dados", "base_unificada_fiscal_2026-08-05.xlsx — 6.898 registros (522 chamados + 6.376 conversas) com a classificação de cada um."],
 ["Relatórios", "analise_chamados_fiscais_2026-08-04.md e analise_balcao_conversas_2026-08-05.md"],
 ["Período analisado", "2026-05-01 a 2026-08-05"],
 ["Campos do Bitrix usados nos cálculos", "createdDate = data de abertura do chamado (preenchido em 100% dos 1.227). · dateStart = data de início, quando o time começou a desenvolver (799 chamados; falta em 35%). · changedDate = data da última modificação do chamado; nos concluídos coincide com o fechamento em 555 de 583 casos, e quando difere a mediana é de 1,2 dia. · closedDate = data de fechamento (872 chamados; bate 100% com status = concluída). · Não usamos timeSpentInLogs (tempo apontado), que está vazio em todos os chamados porque o controle de tempo do Bitrix está desligado — por isso todas as métricas são de calendário, nunca de esforço real."],
 ["Revisão", "Os planos de ação desta planilha foram revisados e editados por Lucas em 05/08/2026. Essa versão revisada é a fonte da verdade (analysis/acoes_user.json) e tem precedência sobre o texto gerado automaticamente."],
 ["Gerado em", "2026-08-05"],
]
for r in leg:
    ws3.append(r)
ws3.column_dimensions["A"].width = 22
ws3.column_dimensions["B"].width = 130
for r in ws3.iter_rows():
    r[0].font = Font(bold=True)
    r[1].alignment = Alignment(wrap_text=True, vertical="top")

wb.save(OUT)
print(f"Salvo: {OUT}")
n_ac = sum(r["PlanoDeAcao"].count("\n") + 1 for r in rows)
print(f"{len(rows)} subgrupos, {n_ac} ações no total")
